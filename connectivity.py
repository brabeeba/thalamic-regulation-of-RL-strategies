from util import load_dict, save_dict, relu
import numpy as np
from torch.utils.data import DataLoader, Dataset
import torch.nn as nn
import torch.nn.functional as F
import torch
import os
import matplotlib
import matplotlib.pyplot as plt
import scipy
import scikit_posthocs as sp
import seaborn as sns
import pandas as pd
from sklearn.linear_model import ridge_regression
from openpyxl import load_workbook

# def ridge_regression(X, Y, mask):
# 	input_dim = X.shape[1]
# 	num = X.shape[0]
# 	output_dim = Y.shape[1]

# 	lower, upper = -(np.sqrt(6.0) / np.sqrt(input_dim + output_dim)), (np.sqrt(6.0) / np.sqrt(input_dim + output_dim))
# 	W = lower + np.random.rand(input_dim, output_dim) * (upper - lower)
# 	W = W * mask
# 	iteration = 2000
# 	lr = 0.001

# 	for t in range(iteration):
# 		loss = np.mean(np.sum((X.dot(W) - Y) * (X.dot(W) - Y), axis = 1)) + np.sum(W * W)

# 		if t % 200 == 0:
# 			print("loss {} at step {}".format(loss, t))

# 		W -= lr * (X.T.dot((X.dot(W) - Y)) / num + 2 *  W)
# 		W = W * mask

# 	return W * mask




matplotlib.rcParams['pdf.fonttype'] = 42
matplotlib.rcParams['ps.fonttype'] = 42


sns.set_theme(context = "paper", style = "ticks")
#sns.set_theme("ticks", palette=None, context = "paper")
palette = sns.color_palette()
palette_alt = sns.color_palette("Set2")
palette_gradient = sns.color_palette("rocket")


data = load_dict("experiment{}_data".format(1))
agents = list(data["action"].keys())
switch = len(data["switch"][agents[0]][0])
max_trial = len(data["action"][agents[0]][0])
max_episode = len(data["action"][agents[0]])
print(max_episode)
block_size = 25
stimuli_num = 2
class_num = 2
context_num = 2
scan_num = 20

mb_idx = []
mf_idx = []
context_score = []
for i in range(max_episode):
	md = np.vstack([ x["MD"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial, scan_num, 2)[:, :, -1, :]
	df = md[i, 25:, 1]
	context_score.append(np.mean(df))

mb_idx = np.arange(max_episode)[scipy.stats.zscore(context_score) > 0]
mf_idx = np.arange(max_episode)[scipy.stats.zscore(context_score) < 0]
		

# dfs = pd.read_excel("MF_vs_MB1.xlsx", sheet_name=None)

# ind_df = np.concatenate([[[0, 0]],dfs["Sheet1"].to_numpy()], axis = 0)[:, 1]
# mb_idx = (ind_df == 0)
# mf_idx = (ind_df == 1)

# print(np.sum(mb_idx), len(mb_idx))
# correct = 0
# for i, v in enumerate(ind_df):
# 	if v == 0 and i in mb_idx:
# 		correct += 1
# 	if v == 1 and i in mf_idx:
# 		correct += 1
# decoded_accuracy = correct / 500
# print(decoded_accuracy)




print(data["histogram"][agents[0]][0]["MD"].shape)


pfc = np.vstack([ x["PFC"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial, scan_num, stimuli_num * class_num * 2)
alm = np.vstack([ x["ALM"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial, scan_num, 2)

# stimuli = data["stimuli"][agents[0]]
# alm_inputs = np.zeros((max_episode, max_trial, scan_num, stimuli_num))

# for i in range(max_episode):
# 	for j in range(max_trial):
# 		for k in range(scan_num):
# 			alm_inputs[i, j, k, stimuli[i, j]] = 1





#alm = np.vstack([ x["ALM"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial , scan_num, stimuli_num)
PFC = np.concatenate([pfc, alm], axis = 3)
OFC = np.vstack([ x["BG"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial , scan_num, context_num * class_num )
MD = np.vstack([ x["MD"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial , scan_num, 2) 


stimuli = data["stimuli"][agents[0]]
action = data["action"][agents[0]]
reward = data["action"][agents[0]]

prob = np.vstack([ x["ALM/BG"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial , scan_num, context_num , stimuli_num, class_num)[:, :, -1, :, :, :]
lr = np.vstack([ x["VIP"] - x["PV"] for x in data["histogram"][agents[0]]]).reshape(max_episode, max_trial , scan_num, context_num)[:, :, -1, :]
lr = relu(2 /  (1 + np.exp( - 2*(lr)))-1)
print(prob.shape)

value = np.zeros((max_episode, max_trial, context_num))
for i in range(max_episode):
	for j in range(max_trial):
		for k in range(context_num):
			value[i, j, k] = prob[i, j, k, stimuli[i, j], action[i, j]]



rpe = np.expand_dims(reward, axis = 2) - value  

rpe = np.concatenate([np.zeros((max_episode, 1, 2)), rpe[:, :-1, :]], axis = 1)

df = np.concatenate([PFC, OFC, MD], axis = 3)

# mask = np.ones((18, 18))
# mask[14, 16:18] = np.zeros(2)
# mask[15, 16:18] = np.zeros(2)
# mask[16, [14, 15, 17]] = np.zeros(3)
# mask[17, 14:17] = np.zeros(3)
# mask[16, :10] = np.zeros(10)
# mask[17, :10] = np.zeros(10)
# mask[14, :10] = np.zeros(10)
# mask[15, :10] = np.zeros(10)
# mask[8, -4:] = np.zeros(4)
# mask[9, -4:] = np.zeros(4)
mask = np.ones((16, 16))
fig, ax = plt.subplots()
ax.matshow(mask)
fig.savefig("mask.pdf")
plt.close()


df_x = df.copy()[:, :, 0, :]
df_x = np.concatenate([df_x, rpe ], axis = 2)

df_y = df.copy()[:, :, -1, :]
df_y = np.concatenate([df_y, rpe ], axis = 2)

steady_df_x = df_x[:, 15:25, :]
steady_df_y = df_y[:, 15:25, :]
steady_df_x = steady_df_x.reshape(steady_df_x.shape[0] , steady_df_x.shape[1], steady_df_x.shape[2])
steady_df_y = steady_df_y.reshape(steady_df_y.shape[0] , steady_df_y.shape[1], steady_df_y.shape[2])


mb_df_x = df_x[mb_idx, 25:35, :]
mb_df_y = df_y[mb_idx, 25:35, :]
mb_df_x = mb_df_x.reshape(mb_df_x.shape[0] , mb_df_x.shape[1], mb_df_x.shape[2])
mb_df_y = mb_df_y.reshape(mb_df_y.shape[0] , mb_df_y.shape[1], mb_df_y.shape[2])

mf_df_x = df_x[mf_idx, 25:35, :]
mf_df_y = df_y[mf_idx, 25:35, :]
mf_df_x = mf_df_x.reshape(mf_df_x.shape[0] , mf_df_x.shape[1], mf_df_x.shape[2])
mf_df_y = mf_df_y.reshape(mf_df_y.shape[0] , mf_df_y.shape[1], mf_df_y.shape[2])

label = ["SS", "MF", "MB"]

steady_cortico = []
coef = ridge_regression(steady_df_x.reshape(steady_df_x.shape[0] * steady_df_x.shape[1], steady_df_x.shape[2]), steady_df_y.reshape(steady_df_y.shape[0] * steady_df_y.shape[1], steady_df_y.shape[2]), alpha = 1)
for i in range(steady_df_x.shape[0]):
	PFC_to_OFC = coef[10:14, :10].dot(steady_df_x[i, :, :10].T)
	OFC_to_PFC = coef[:10, 10:14].dot(steady_df_x[i, :, 10:14].T)
	

	cortico = np.concatenate([OFC_to_PFC, PFC_to_OFC], axis = 0) * (steady_df_y[i, :, :14]/ np.linalg.norm(steady_df_y[i, :, :14], axis = 1, keepdims=True)).T

	steady_cortico.append(np.sum(cortico))

mb_cortico = []
coef = ridge_regression(mb_df_x.reshape(mb_df_x.shape[0] * mb_df_x.shape[1], mb_df_x.shape[2]), mb_df_y.reshape(mb_df_y.shape[0] * mb_df_y.shape[1], mb_df_y.shape[2]), alpha = 1)

for i in range(mb_df_x.shape[0]):
	PFC_to_OFC = coef[10:14, :10].dot(mb_df_x[i, :, :10].T)
	OFC_to_PFC = coef[:10, 10:14].dot(mb_df_x[i, :, 10:14].T)

	
	cortico = np.concatenate([OFC_to_PFC, PFC_to_OFC], axis = 0) * (mb_df_y[i, :, :14]/ np.linalg.norm(mb_df_y[i, :, :14], axis = 1, keepdims=True)).T

	mb_cortico.append(np.sum(cortico))

mf_cortico = []
coef = ridge_regression(mf_df_x.reshape(mf_df_x.shape[0] * mf_df_x.shape[1], mf_df_x.shape[2]), mf_df_y.reshape(mf_df_y.shape[0] * mf_df_y.shape[1], mf_df_y.shape[2]), alpha = 1)

for i in range(mf_df_x.shape[0]):
	

	PFC_to_OFC = coef[10:14, :10].dot(mf_df_x[i, :, :10].T)
	OFC_to_PFC = coef[:10, 10:14].dot(mf_df_x[i, :, 10:14].T)
	

	cortico = np.concatenate([OFC_to_PFC, PFC_to_OFC], axis = 0) * (mf_df_y[i, :, :14]/ np.linalg.norm(mf_df_y[i, :, :14], axis = 1, keepdims=True)).T

	mf_cortico.append(np.sum(cortico))

df_list = []

ratio = 0.8
fig, ax = plt.subplots()
fig.set_figwidth(5.5 * ratio)
fig.set_figheight(4.8 * ratio)

cortico_data = [steady_cortico, mf_cortico, mb_cortico]

label = ["SS", "MF", "MB"]
t = range(3)

df1 = pd.DataFrame()
df1['cortico_connectivity'] = cortico_data[0]
df1['type'] = label[0]

df2 = pd.DataFrame()
df2['cortico_connectivity'] = cortico_data[1]
df2['type'] = label[1]

df3 = pd.DataFrame()
df3['cortico_connectivity'] = cortico_data[2]
df3['type'] = label[2]

df_list.append(pd.concat([df1, df2, df3], ignore_index = True))

error_bar =  [ np.std(data) / np.sqrt(len(data)) for data in cortico_data]
mean = [ np.mean(data) for data in cortico_data]
plt.errorbar(t,mean, yerr = error_bar, c = palette[0], capsize=0)
plt.plot(t, mean, c = palette[0])

z, p = scipy.stats.kruskal(cortico_data[0], cortico_data[1], cortico_data[2])

print("kruskal wallis p value for cortico connectivity is {}".format(p))

p = sp.posthoc_dunn(cortico_data, p_adjust='bonferroni')
print(p)


print("Cortico connectivity for {} = {}, sem = {}".format(label[0], np.mean(cortico_data[0]), np.std(cortico_data[0]) / np.sqrt(len(cortico_data[0]))))
print("Cortico connectivity for {} = {}, sem = {}".format(label[1],  np.mean(cortico_data[1]), np.std(cortico_data[1]) / np.sqrt(len(cortico_data[1]))))
print("Cortico connectivity for {} = {}, sem = {}".format(label[2],  np.mean(cortico_data[2]), np.std(cortico_data[2]) / np.sqrt(len(cortico_data[2]))))


	
plt.legend(loc="upper left", frameon=False)
plt.ylabel("Cortical connectivity")
ax.set_xticks(t)
ax.set_xticklabels(label) 
sns.despine()
plt.savefig("fig/cortical_connectivity_box.pdf", transparent = True)
plt.close()


steady_thalamo = []
coef = ridge_regression(steady_df_x.reshape(steady_df_x.shape[0] * steady_df_x.shape[1], steady_df_x.shape[2]), steady_df_y.reshape(steady_df_y.shape[0] * steady_df_y.shape[1], steady_df_y.shape[2]), alpha = 1)
for i in range(steady_df_x.shape[0]):
	MD_to_OFC = coef[10:14, -4:].dot(steady_df_x[i, :, -4:].T)
	MD_to_PFC = coef[:10, -4:].dot(steady_df_x[i, :, -4:].T)
	
	#print(steady_df_y[i, :, idx])
	
	thalamo = np.concatenate([ MD_to_PFC, MD_to_OFC], axis = 0) * (steady_df_y[i, :, :14]/ np.linalg.norm(steady_df_y[i, :, :14], axis = 1, keepdims=True)).T

	steady_thalamo.append(np.sum(thalamo))
	

mb_thalamo = []
coef = ridge_regression(mb_df_x.reshape(mb_df_x.shape[0] * mb_df_x.shape[1], mb_df_x.shape[2]), mb_df_y.reshape(mb_df_y.shape[0] * mb_df_y.shape[1], mb_df_y.shape[2]), alpha = 1)

for i in range(mb_df_x.shape[0]):
	MD_to_OFC = coef[10:14, -4:].dot(mb_df_x[i, :, -4:].T)
	MD_to_PFC = coef[:10, -4:].dot(mb_df_x[i, :, -4:].T)


	thalamo = np.concatenate([ MD_to_PFC, MD_to_OFC], axis = 0) * (mb_df_y[i, :, :14]/ np.linalg.norm(mb_df_y[i, :, :14], axis = 1, keepdims=True)).T

	mb_thalamo.append(np.sum(thalamo))
mf_thalamo = []
coef = ridge_regression(mf_df_x.reshape(mf_df_x.shape[0] * mf_df_x.shape[1], mf_df_x.shape[2]), mf_df_y.reshape(mf_df_y.shape[0] * mf_df_y.shape[1], mf_df_y.shape[2]), alpha = 1)

for i in range(mf_df_x.shape[0]):
	MD_to_OFC = coef[10:14, -4:].dot(mf_df_x[i, :, -4:].T)
	MD_to_PFC = coef[:10, -4:].dot(mf_df_x[i, :, -4:].T)

	
	thalamo = np.concatenate([ MD_to_PFC, MD_to_OFC], axis = 0) * (mf_df_y[i, :, :14]/ np.linalg.norm(mf_df_y[i, :, :14], axis = 1, keepdims=True)).T

	mf_thalamo.append(np.sum(thalamo))

ratio = 0.8
fig, ax = plt.subplots()
fig.set_figwidth(5.5 * ratio)
fig.set_figheight(4.8 * ratio)

thalamo_data = [steady_thalamo, mf_thalamo, mb_thalamo]
label = ["SS", "MF", "MB"]
t = range(3)


df1 = pd.DataFrame()
df1['thalamocortical_connectivity'] = thalamo_data[0]
df1['type'] = label[0]

df2 = pd.DataFrame()
df2['thalamocortical_connectivity'] = thalamo_data[1]
df2['type'] = label[1]

df3 = pd.DataFrame()
df3['thalamocortical_connectivity'] = thalamo_data[2]
df3['type'] = label[2]

df_list.append(pd.concat([df1, df2, df3], ignore_index = True))


error_bar =  [ np.std(data) / np.sqrt(len(data)) for data in thalamo_data]
mean = [ np.mean(data) for data in thalamo_data]
plt.errorbar(t,mean, yerr = error_bar, c = palette[0], capsize=0)
plt.plot(t, mean, c = palette[0])



z, p = scipy.stats.kruskal(thalamo_data[0], thalamo_data[1], thalamo_data[2])
print("kruskal wallis p value for cortico connectivity is {}".format(p))

p = sp.posthoc_dunn(thalamo_data, p_adjust='bonferroni')
print(p)


print("Thalamocortical connectivity for {} = {}, sem = {}".format(label[0], np.mean(thalamo_data[0]), np.std(thalamo_data[0]) / np.sqrt(len(thalamo_data[0]))))
print("Thalamocortical connectivity for {} = {}, sem = {}".format(label[1],  np.mean(thalamo_data[1]), np.std(thalamo_data[1]) / np.sqrt(len(thalamo_data[1]))))
print("Thalamocortical connectivity for {} = {}, sem = {}".format(label[2],  np.mean(thalamo_data[2]), np.std(thalamo_data[2]) / np.sqrt(len(thalamo_data[2]))))


	
plt.legend(loc="upper left", frameon=False)
plt.ylabel("Thalamocortical connectivity")
ax.set_xticks(t)
ax.set_xticklabels(label) 
sns.despine()
plt.savefig("fig/thalamocortical_connectivity_box.pdf", transparent = True)
plt.close()


fig, ax = plt.subplots(3)

coef = ridge_regression(steady_df_x.reshape(steady_df_x.shape[0] * steady_df_x.shape[1], steady_df_x.shape[2]), steady_df_y.reshape(steady_df_y.shape[0] * steady_df_y.shape[1], steady_df_y.shape[2]), alpha = 1)


pfc_norm = np.mean(df_x[:, :, :10], axis = (0,1)) 
md_norm = np.mean(df_x[:, :, -4:-2], axis = (0, 1))
print( pfc_norm, md_norm)
print("hi", np.min(np.concatenate([coef[10:14, :10] * pfc_norm , coef[10:14, -4:-2]  * md_norm]  , axis = 1)), np.max(np.concatenate([coef[10:14, :10] * pfc_norm , coef[10:14, -4:-2] * md_norm]  , axis = 1)))


a = ax[0].matshow(np.concatenate([coef[10:14, :10] * pfc_norm , coef[10:14, -4:-2] * md_norm]  , axis = 1), vmin = -0.2, vmax = 0.2, cmap='RdBu')
fig.colorbar(a, orientation='horizontal')
coef = ridge_regression(mb_df_x.reshape(mb_df_x.shape[0] * mb_df_x.shape[1], mb_df_x.shape[2]), mb_df_y.reshape(mb_df_y.shape[0] * mb_df_y.shape[1], mb_df_y.shape[2]), alpha = 1)

a =  ax[2].matshow(np.concatenate([coef[10:14, :10] * pfc_norm , coef[10:14, -4:-2] * md_norm]  , axis = 1), vmin = -0.2, vmax = 0.2, cmap='RdBu')
fig.colorbar(a, orientation='horizontal')
coef = ridge_regression(mf_df_x.reshape(mf_df_x.shape[0] * mf_df_x.shape[1], mf_df_x.shape[2]), mf_df_y.reshape(mf_df_y.shape[0] * mf_df_y.shape[1], mf_df_y.shape[2]), alpha = 1)


a =  ax[1].matshow(np.concatenate([coef[10:14, :10] * pfc_norm , coef[10:14, -4:-2] * md_norm]   , axis = 1), vmin = -0.2, vmax = 0.2, cmap='RdBu')
fig.colorbar(a, orientation='horizontal')
plt.savefig("fig/thalamocotical_connectivity_matrix.pdf", transparent = True)

fig, ax = plt.subplots(3)

norm = np.mean(df_x[:-4, :], axis = (0, 1))
coef = ridge_regression(steady_df_x.reshape(steady_df_x.shape[0] * steady_df_x.shape[1], steady_df_x.shape[2]), steady_df_y.reshape(steady_df_y.shape[0] * steady_df_y.shape[1], steady_df_y.shape[2]), alpha = 1)
print(np.min(coef[:-4, :] * norm), np.max(coef[:-4, :] * norm))
a = ax[0].matshow(coef[:-4, :] * norm, vmin = -0.4, vmax = 0.4, cmap='RdBu')
fig.colorbar(a)

coef = ridge_regression(mb_df_x.reshape(mb_df_x.shape[0] * mb_df_x.shape[1], mb_df_x.shape[2]), mb_df_y.reshape(mb_df_y.shape[0] * mb_df_y.shape[1], mb_df_y.shape[2]), alpha = 1)

a = ax[2].matshow(coef[:-4, :] * norm, vmin = -0.4, vmax = 0.4, cmap='RdBu')

coef = ridge_regression(mf_df_x.reshape(mf_df_x.shape[0] * mf_df_x.shape[1], mf_df_x.shape[2]), mf_df_y.reshape(mf_df_y.shape[0] * mf_df_y.shape[1], mf_df_y.shape[2]), alpha = 1)


a = ax[1].matshow(coef[:-4, :] * norm, vmin = -0.4, vmax = 0.4,  cmap='RdBu')


plt.savefig("fig/cortico_connectivity_matrix.pdf", transparent = True)



file_path = "output.xlsx"

# Load existing workbook
book = load_workbook(file_path)

if "6e-l" in book.sheetnames:
	std = book["6e-l"]
	book.remove(std)
if "6e-r" in book.sheetnames:
	std = book["6e-r"]
	book.remove(std)

sheet1 = book.create_sheet("6e-l")
for r_idx, row in enumerate(df1.itertuples(index=False), start=2):
	for c_idx, value in enumerate(row, start=1):
		sheet1.cell(row=r_idx, column=c_idx, value=value)
# Write header
for c_idx, col_name in enumerate(df1.columns, start=1):
	sheet1.cell(row=1, column=c_idx, value=col_name)

# Add second sheet
sheet2 = book.create_sheet("6e-r")
for r_idx, row in enumerate(df2.itertuples(index=False), start=2):
	for c_idx, value in enumerate(row, start=1):
		sheet2.cell(row=r_idx, column=c_idx, value=value)
for c_idx, col_name in enumerate(df2.columns, start=1):
	sheet2.cell(row=1, column=c_idx, value=col_name)

# Save the workbook
book.save(file_path)


